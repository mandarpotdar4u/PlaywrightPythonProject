import os

import openpyxl
from config import settings


# ---> Reading & Fetching values from Excel sheet
def get_valid_login_data_from_excel():
    # base_path = os.path.dirname(__file__)
    # config_path = os.path.join(base_path, '../data/testfile.xlsx')
    config_path = settings.excel_config_path
    final_list = []
    workbook = openpyxl.load_workbook(config_path)
    sheet = workbook[settings.excel_sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
        return final_list


# def get_invalid_login_data_from_excel(path, sheet_name):
# final_list = []
# workbook = openpyxl.load_workbook(path)
# sheet = workbook[sheet_name]
# total_rows = sheet.max_row
# total_cols = sheet.max_column
#
# for r in range(3, total_rows + 1):
#     row_list = []
#     for c in range(1, total_cols + 1):
#         row_list.append(sheet.cell(row=r, column=c).value)
#     final_list.append(row_list)
#     return final_list

def get_invalid_login_data_from_excel():
    base_path = os.path.dirname(__file__)
    config_path = settings.excel_config_path
    final_list = []
    workbook = openpyxl.load_workbook(config_path)
    sheet = workbook[settings.excel_sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(3, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
        return final_list


def get_valid_Page_Title_assertion_from_excel():
    base_path = os.path.dirname(__file__)
    # config_path = os.path.join(base_path, '../data/testfile.xlsx')
    config_path = settings.excel_config_path
    final_list = []
    workbook = openpyxl.load_workbook(config_path)
    sheet = workbook[settings.excel_sheet_name]
    # final_list = []
    # workbook = openpyxl.load_workbook(config_path)
    # sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(5, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
        return final_list


def get_invalid_Page_Title_assertion_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(7, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
        return final_list


def get_invalid_Login_Credentials_wARNING_assertion_from_excel(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(3, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)
        return final_list
